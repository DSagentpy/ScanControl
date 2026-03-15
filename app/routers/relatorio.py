from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import func
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    HRFlowable,
    PageTemplate,
    Frame
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle

from datetime import datetime
from io import BytesIO

from app.database import get_db
from app import models

router = APIRouter()


# 🔹 Header e Footer estilo ERP
def header_footer(canvas_obj, doc):
    canvas_obj.saveState()
    width, height = A4

    # Cabeçalho
    canvas_obj.setFont("Helvetica-Bold", 10)
    canvas_obj.drawString(2 * cm, height - 1.5 * cm,
                          "SISTEMA DE CONTROLE LOGÍSTICO")

    canvas_obj.setFont("Helvetica", 8)
    canvas_obj.drawRightString(
        width - 2 * cm,
        height - 1.5 * cm,
        f"Emitido em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    )

    # Rodapé
    canvas_obj.setFont("Helvetica", 8)
    canvas_obj.drawCentredString(width / 2, 1.2 * cm,
                                 f"Página {doc.page}")

    canvas_obj.restoreState()


# 🔥 Endpoint PDF
@router.get("/relatorio/pdf")
def gerar_pdf(
    sessao_id: int = Query(None, description="ID da sessão (opcional, usa última se não informado)"),
    titulo: str = Query("Relatório de Conferência"),
    db: Session = Depends(get_db)
):

    if sessao_id:
        sessao = db.query(models.Sessao).filter(
            models.Sessao.id == sessao_id
        ).first()
    else:
        sessao = db.query(models.Sessao).order_by(
            models.Sessao.id.desc()
        ).first()

    if not sessao:
        raise HTTPException(status_code=404,
                            detail="Nenhuma sessão encontrada")

    resultados = db.query(
        models.Produto.codigo,
        models.Produto.nome,
        models.Produto.descricao,
        models.Produto.codigo_barra,
        func.sum(models.Recebimento.quantidade).label("total")
    ).join(
        models.Recebimento,
        models.Produto.id == models.Recebimento.produto_id
    ).filter(
        models.Recebimento.sessao_id == sessao.id
    ).group_by(
        models.Produto.codigo,
        models.Produto.nome,
        models.Produto.descricao,
        models.Produto.codigo_barra
    ).order_by(
        models.Produto.nome
    ).all()

    if not resultados:
        raise HTTPException(status_code=400,
                            detail="Nenhum produto na sessão")

    buffer = BytesIO()

    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=3 * cm,
        bottomMargin=2 * cm
    )

    elements = []
    styles = getSampleStyleSheet()

    # 🔹 Título
    titulo_style = ParagraphStyle(
        'TituloERP',
        parent=styles['Title'],
        fontSize=16,
        textColor=colors.HexColor("#1F4E79")
    )

    nome_sessao_pdf = sessao.nome or f"Sessão {sessao.id}"
    elements.append(Paragraph(f"{titulo} — {nome_sessao_pdf}", titulo_style))
    elements.append(Spacer(1, 12))
    elements.append(HRFlowable(width="100%", thickness=1,
                               color=colors.grey))
    elements.append(Spacer(1, 20))

    # 🔹 Informações da sessão
    elements.append(Paragraph(
        f"<b>Sessão #{sessao.id}</b>{(' — ' + sessao.nome) if sessao.nome else ''}",
        styles["Normal"]
    ))
    elements.append(Spacer(1, 12))

    # 🔹 Tabela
    data = [["Código", "Nome", "Código de Barras", "Qtd"]]

    total_geral = 0
    total_itens = len(resultados)

    for r in resultados:
        total_geral += r.total
        data.append([
            r.codigo,
            r.nome or r.descricao or "",
            r.codigo_barra,
            str(r.total)
        ])

    table = Table(
        data,
        colWidths=[3 * cm, 6.5 * cm, 4.5 * cm, 2 * cm],
        repeatRows=1
    )

    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0),
         colors.HexColor("#D9D9D9")),
        ('FONTNAME', (0, 0), (-1, 0),
         'Helvetica-Bold'),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('GRID', (0, 0), (-1, -1),
         0.3, colors.grey),
        ('ALIGN', (-1, 1), (-1, -1),
         'CENTER'),
    ]))

    elements.append(table)
    elements.append(Spacer(1, 25))

    # 🔹 Resumo executivo
    resumo_data = [
        ["Total de Produtos Diferentes:",
         str(total_itens)],
        ["Quantidade Total Recebida:",
         str(total_geral)]
    ]

    resumo_table = Table(resumo_data,
                         colWidths=[8 * cm, 3 * cm])

    resumo_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, -1),
         colors.HexColor("#F2F2F2")),
        ('GRID', (0, 0), (-1, -1),
         0.5, colors.grey),
        ('FONTNAME', (0, 0), (-1, -1),
         'Helvetica-Bold'),
        ('ALIGN', (-1, 0), (-1, -1),
         'CENTER')
    ]))

    elements.append(resumo_table)

    # 🔹 Template com header/footer
    frame = Frame(
        doc.leftMargin,
        doc.bottomMargin,
        doc.width,
        doc.height,
        id='normal'
    )

    template = PageTemplate(
        id='ERPTemplate',
        frames=frame,
        onPage=header_footer
    )

    doc.addPageTemplates([template])
    doc.build(elements)

    buffer.seek(0)

    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition":
            "attachment; filename=relatorio_erp.pdf"
        }
    )


# ─── Endpoint Excel ───────────────────────────────────────────────────────────
@router.get("/relatorio/excel")
def gerar_excel(
    sessao_id: int = Query(None, description="ID da sessão (usa última se omitido)"),
    db: Session = Depends(get_db)
):
    if sessao_id:
        sessao = db.query(models.Sessao).filter(
            models.Sessao.id == sessao_id
        ).first()
    else:
        sessao = db.query(models.Sessao).order_by(
            models.Sessao.id.desc()
        ).first()

    if not sessao:
        raise HTTPException(status_code=404, detail="Nenhuma sessão encontrada")

    resultados = db.query(
        models.Produto.codigo,
        models.Produto.nome,
        models.Produto.descricao,
        models.Produto.codigo_barra,
        func.sum(models.Recebimento.quantidade).label("total")
    ).join(
        models.Recebimento,
        models.Produto.id == models.Recebimento.produto_id
    ).filter(
        models.Recebimento.sessao_id == sessao.id
    ).group_by(
        models.Produto.codigo,
        models.Produto.nome,
        models.Produto.descricao,
        models.Produto.codigo_barra
    ).order_by(
        models.Produto.nome
    ).all()

    if not resultados:
        raise HTTPException(status_code=400, detail="Nenhum produto na sessão")

    # ── Criar planilha ────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Sessão {sessao.id}"

    # Estilos
    cor_cabecalho = "1F4E79"
    cor_linha_par = "D6E4F0"
    cor_total = "E8F5E9"

    estilo_borda = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ── Título ────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    titulo_cell = ws["A1"]
    nome_sessao = sessao.nome or f"Sessão {sessao.id}"
    titulo_cell.value = f"RELATÓRIO DE CONFERÊNCIA — {nome_sessao.upper()}"
    titulo_cell.font = Font(bold=True, size=14, color="FFFFFF")
    titulo_cell.fill = PatternFill("solid", fgColor=cor_cabecalho)
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Subtítulo ─────────────────────────────────────────────────────────────
    ws.merge_cells("A2:E2")
    sub = ws["A2"]
    data_emissao = datetime.now().strftime("%d/%m/%Y %H:%M")
    sub.value = f"Emitido em: {data_emissao}  |  Sessão #{sessao.id}"
    sub.font = Font(italic=True, size=10, color="555555")
    sub.alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # ── Cabeçalho da tabela ───────────────────────────────────────────────────
    cabecalhos = ["Código", "Nome", "Descrição", "Código de Barras", "Quantidade"]
    for col, cab in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=4, column=col, value=cab)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill("solid", fgColor=cor_cabecalho)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = estilo_borda
    ws.row_dimensions[4].height = 20

    # ── Dados ─────────────────────────────────────────────────────────────────
    total_geral = 0
    for i, r in enumerate(resultados, start=5):
        total_geral += r.total
        linha = [r.codigo, r.nome, r.descricao or "", r.codigo_barra, r.total]
        fill = PatternFill("solid", fgColor=cor_linha_par) if i % 2 == 0 else None

        for col, valor in enumerate(linha, start=1):
            cell = ws.cell(row=i, column=col, value=valor)
            cell.border = estilo_borda
            cell.alignment = Alignment(
                horizontal="center" if col in (1, 5) else "left",
                vertical="center"
            )
            if fill:
                cell.fill = fill

    # ── Linha de totais ───────────────────────────────────────────────────────
    linha_total = len(resultados) + 5
    ws.merge_cells(f"A{linha_total}:D{linha_total}")
    celula_label = ws[f"A{linha_total}"]
    celula_label.value = f"TOTAL — {len(resultados)} produto(s) diferente(s)"
    celula_label.font = Font(bold=True, size=11)
    celula_label.fill = PatternFill("solid", fgColor=cor_total)
    celula_label.alignment = Alignment(horizontal="right", vertical="center")
    celula_label.border = estilo_borda

    celula_total = ws[f"E{linha_total}"]
    celula_total.value = total_geral
    celula_total.font = Font(bold=True, size=12)
    celula_total.fill = PatternFill("solid", fgColor=cor_total)
    celula_total.alignment = Alignment(horizontal="center", vertical="center")
    celula_total.border = estilo_borda
    ws.row_dimensions[linha_total].height = 22

    # ── Largura das colunas ───────────────────────────────────────────────────
    larguras = [12, 30, 30, 18, 12]
    for col, largura in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = largura

    # ── Salvar e retornar ─────────────────────────────────────────────────────
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    nome_arquivo = f"conferencia_sessao_{sessao.id}.xlsx"
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
    )