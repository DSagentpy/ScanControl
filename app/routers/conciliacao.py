"""
Router de Auditoria de Inventário
Etapa 2 do pipeline: Conciliação Físico vs. Sistêmico + KPIs automáticos
"""
import io
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

from app.database import get_db
from app import models

router = APIRouter()


# ── Constantes de status ──────────────────────────────────────────────────────

STATUS_LABEL = {
    "OK":            "OK ✅",
    "DIVERGENTE":    "Divergente ⚠️",
    "CRITICO":       "Crítico 🚨",
    "FANTASMA":      "Fantasma 👻",
    "NAO_SISTEMICO": "Não Sistêmico 🔵",
    "RUPTURA":       "Ruptura 🔴",
}

STATUS_COR_EXCEL = {
    "OK":            "C8E6C9",  # verde claro
    "DIVERGENTE":    "FFF9C4",  # amarelo claro
    "CRITICO":       "FFCDD2",  # vermelho claro
    "FANTASMA":      "E1BEE7",  # roxo claro
    "NAO_SISTEMICO": "BBDEFB",  # azul claro
    "RUPTURA":       "FFE0B2",  # laranja claro
}


# ── Leitura e normalização do arquivo ERP ────────────────────────────────────

def _ler_arquivo_erp(arquivo: UploadFile) -> pd.DataFrame:
    """
    Lê CSV ou Excel do ERP e normaliza colunas.
    Colunas esperadas (nomes flexíveis):
      - codigo_barra  (obrigatório)
      - quantidade    (obrigatório)
      - estoque_minimo (opcional)
      - nome          (opcional)
    """
    conteudo = arquivo.file.read()
    nome = (arquivo.filename or "").lower()

    if nome.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(conteudo), encoding="utf-8",
                         sep=None, engine="python", dtype=str)
    elif nome.endswith((".xlsx", ".xls")):
        df = pd.read_excel(io.BytesIO(conteudo), dtype=str)
    else:
        raise ValueError("Formato não suportado. Use .csv, .xlsx ou .xls")

    # Normalizar nomes
    df.columns = [str(c).strip().lower()
                  .replace(" ", "_")
                  .replace("-", "_")
                  for c in df.columns]

    aliases = {
        "cod_barra": "codigo_barra", "codigobarra": "codigo_barra",
        "ean": "codigo_barra", "barcode": "codigo_barra",
        "cod": "codigo_barra", "codigo": "codigo_barra",
        "qtd": "quantidade", "qtd_sistemica": "quantidade",
        "quantidade_sistemica": "quantidade", "qty": "quantidade",
        "estoque": "quantidade", "stock": "quantidade",
        "qtd_atual": "quantidade",
        "min": "estoque_minimo", "minimo": "estoque_minimo",
        "estoque_min": "estoque_minimo", "estq_min": "estoque_minimo",
        "produto": "nome", "descricao": "nome", "item": "nome",
    }
    df.rename(columns=aliases, inplace=True)

    if "codigo_barra" not in df.columns:
        raise ValueError(
            "Coluna 'codigo_barra' não encontrada. "
            "Renomeie a coluna com o código de barras para 'codigo_barra'."
        )
    if "quantidade" not in df.columns:
        raise ValueError(
            "Coluna 'quantidade' não encontrada. "
            "Renomeie a coluna com a quantidade sistêmica para 'quantidade'."
        )

    df["codigo_barra"] = df["codigo_barra"].astype(str).str.strip()
    df["quantidade"] = pd.to_numeric(
        df["quantidade"], errors="coerce").fillna(0).astype(int)

    if "estoque_minimo" not in df.columns:
        df["estoque_minimo"] = 0
    else:
        df["estoque_minimo"] = pd.to_numeric(
            df["estoque_minimo"], errors="coerce").fillna(0).astype(int)

    if "nome" not in df.columns:
        df["nome"] = ""
    else:
        df["nome"] = df["nome"].fillna("").astype(str).str.strip()

    # Remover linhas sem código
    df = df[df["codigo_barra"].str.len() > 0].copy()
    return df


# ── Lógica de conciliação ─────────────────────────────────────────────────────

def _calcular_status(fisica: int, sistemica: int, minimo: int) -> str:
    if sistemica == 0 and fisica > 0:
        return "NAO_SISTEMICO"
    if fisica == 0 and sistemica > 0:
        return "FANTASMA"
    if fisica == 0 and sistemica == 0:
        return "OK"
    if minimo > 0 and fisica < minimo:
        return "RUPTURA"
    divergencia_pct = abs(fisica - sistemica) / sistemica
    if divergencia_pct == 0:
        return "OK"
    elif divergencia_pct <= 0.05:
        return "DIVERGENTE"
    else:
        return "CRITICO"


def _calcular_acuracidade(fisica: int, sistemica: int) -> float:
    if sistemica == 0:
        return 100.0 if fisica == 0 else 0.0
    return round(min(fisica, sistemica) / sistemica * 100, 2)


def _conciliar(df_fisico: pd.DataFrame, df_erp: pd.DataFrame) -> pd.DataFrame:
    df = pd.merge(df_fisico, df_erp, on="codigo_barra", how="outer")
    df["qtd_fisica"] = df["qtd_fisica"].fillna(0).astype(int)
    df["qtd_sistemica"] = df["qtd_sistemica"].fillna(0).astype(int)
    df["estoque_minimo"] = df["estoque_minimo"].fillna(0).astype(int)
    df["nome"] = df["nome_fisico"].fillna("").where(
        df["nome_fisico"].fillna("") != "",
        df["nome_erp"].fillna("")
    )
    df["divergencia"] = df["qtd_fisica"] - df["qtd_sistemica"]
    df["acuracidade_item"] = df.apply(
        lambda r: _calcular_acuracidade(r["qtd_fisica"], r["qtd_sistemica"]), axis=1
    )
    df["status"] = df.apply(
        lambda r: _calcular_status(r["qtd_fisica"], r["qtd_sistemica"], r["estoque_minimo"]),
        axis=1
    )
    return df


# ── Helpers de resposta ───────────────────────────────────────────────────────

def _auditoria_to_dict(auditoria: models.ConciliacaoAuditoria) -> dict:
    return {
        "id":                   auditoria.id,
        "sessao_id":            auditoria.sessao_id,
        "nome":                 auditoria.nome,
        "criada_em":            auditoria.criada_em.isoformat() if auditoria.criada_em else None,
        "acuracidade_geral":    auditoria.acuracidade_geral,
        "total_itens_fisico":   auditoria.total_itens_fisico,
        "total_itens_sistemico": auditoria.total_itens_sistemico,
        "itens": [
            {
                "id":               i.id,
                "codigo_barra":     i.codigo_barra,
                "nome_produto":     i.nome_produto,
                "qtd_fisica":       i.qtd_fisica,
                "qtd_sistemica":    i.qtd_sistemica,
                "estoque_minimo":   i.estoque_minimo,
                "divergencia":      i.divergencia,
                "acuracidade_item": i.acuracidade_item,
                "status":           i.status,
            }
            for i in sorted(auditoria.itens,
                            key=lambda x: (x.status != "CRITICO",
                                           x.status != "FANTASMA",
                                           x.nome_produto))
        ]
    }


def _calcular_kpis(auditoria: models.ConciliacaoAuditoria) -> dict:
    itens = auditoria.itens
    total = len(itens)
    contagens = {s: 0 for s in STATUS_LABEL}
    for i in itens:
        contagens[i.status] = contagens.get(i.status, 0) + 1

    return {
        "acuracidade_geral":    auditoria.acuracidade_geral or 0,
        "total_produtos":       total,
        "itens_ok":             contagens["OK"],
        "itens_divergentes":    contagens["DIVERGENTE"],
        "itens_criticos":       contagens["CRITICO"],
        "itens_fantasma":       contagens["FANTASMA"],
        "itens_nao_sistemico":  contagens["NAO_SISTEMICO"],
        "itens_em_ruptura":     contagens["RUPTURA"],
        "taxa_ok":              round(contagens["OK"] / total * 100, 1) if total else 0,
        "taxa_divergencia":     round(
            (contagens["DIVERGENTE"] + contagens["CRITICO"]) / total * 100, 1
        ) if total else 0,
        "total_fisico":         auditoria.total_itens_fisico or 0,
        "total_sistemico":      auditoria.total_itens_sistemico or 0,
        "diferenca_total":      (auditoria.total_itens_fisico or 0) -
                                (auditoria.total_itens_sistemico or 0),
    }


# ── Endpoints ─────────────────────────────────────────────────────────────────

@router.post("/conciliacao/executar")
async def executar_conciliacao(
    sessao_id: int = Form(...),
    nome: str = Form(default=""),
    arquivo_erp: UploadFile = File(...),
    db: Session = Depends(get_db)
):
    """
    Executa a conciliação:
    - sessao_id: ID da sessão de scanning (inventário físico)
    - arquivo_erp: CSV ou Excel com o inventário sistêmico
    """
    sessao = db.query(models.Sessao).filter(models.Sessao.id == sessao_id).first()
    if not sessao:
        raise HTTPException(status_code=404, detail="Sessão não encontrada")

    # Ler arquivo ERP
    try:
        df_erp_raw = _ler_arquivo_erp(arquivo_erp)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao processar arquivo: {e}")

    # Inventário físico (escaneado)
    rows = db.query(
        models.Produto.codigo_barra,
        models.Produto.nome,
        func.sum(models.Recebimento.quantidade).label("total")
    ).join(
        models.Recebimento, models.Produto.id == models.Recebimento.produto_id
    ).filter(
        models.Recebimento.sessao_id == sessao_id
    ).group_by(
        models.Produto.codigo_barra, models.Produto.nome
    ).all()

    df_fisico = pd.DataFrame(
        [(r.codigo_barra, r.nome, int(r.total)) for r in rows],
        columns=["codigo_barra", "nome_fisico", "qtd_fisica"]
    )

    df_erp = df_erp_raw.rename(columns={
        "quantidade": "qtd_sistemica",
        "nome": "nome_erp"
    })[["codigo_barra", "nome_erp", "qtd_sistemica", "estoque_minimo"]]

    # Conciliar
    df = _conciliar(df_fisico, df_erp)

    # Acuracidade geral
    total_sist = int(df["qtd_sistemica"].sum())
    total_fisico_valido = int(df.apply(
        lambda r: min(r["qtd_fisica"], r["qtd_sistemica"]), axis=1
    ).sum())
    acuracidade_geral = round(
        total_fisico_valido / total_sist * 100, 2
    ) if total_sist > 0 else 0.0

    # Salvar no banco
    auditoria = models.ConciliacaoAuditoria(
        sessao_id=sessao_id,
        nome=nome.strip() or None,
        acuracidade_geral=acuracidade_geral,
        total_itens_fisico=int(df["qtd_fisica"].sum()),
        total_itens_sistemico=total_sist,
    )
    db.add(auditoria)
    db.flush()

    for _, row in df.iterrows():
        db.add(models.ItemConciliacao(
            auditoria_id=auditoria.id,
            codigo_barra=str(row["codigo_barra"]),
            nome_produto=str(row["nome"]),
            qtd_fisica=int(row["qtd_fisica"]),
            qtd_sistemica=int(row["qtd_sistemica"]),
            estoque_minimo=int(row["estoque_minimo"]),
            divergencia=int(row["divergencia"]),
            acuracidade_item=float(row["acuracidade_item"]),
            status=str(row["status"]),
        ))

    db.commit()
    db.refresh(auditoria)
    return _auditoria_to_dict(auditoria)


@router.get("/conciliacao/historico")
def historico_auditorias(db: Session = Depends(get_db)):
    auditorias = db.query(models.ConciliacaoAuditoria).order_by(
        models.ConciliacaoAuditoria.id.desc()
    ).all()
    return [
        {
            "id":                    a.id,
            "sessao_id":             a.sessao_id,
            "nome":                  a.nome,
            "criada_em":             a.criada_em.isoformat() if a.criada_em else None,
            "acuracidade_geral":     a.acuracidade_geral,
            "total_itens_fisico":    a.total_itens_fisico,
            "total_itens_sistemico": a.total_itens_sistemico,
        }
        for a in auditorias
    ]


@router.get("/conciliacao/{auditoria_id}")
def resultado_auditoria(auditoria_id: int, db: Session = Depends(get_db)):
    auditoria = db.query(models.ConciliacaoAuditoria).filter(
        models.ConciliacaoAuditoria.id == auditoria_id
    ).first()
    if not auditoria:
        raise HTTPException(status_code=404, detail="Auditoria não encontrada")
    return _auditoria_to_dict(auditoria)


@router.get("/conciliacao/{auditoria_id}/kpis")
def kpis_auditoria(auditoria_id: int, db: Session = Depends(get_db)):
    auditoria = db.query(models.ConciliacaoAuditoria).filter(
        models.ConciliacaoAuditoria.id == auditoria_id
    ).first()
    if not auditoria:
        raise HTTPException(status_code=404, detail="Auditoria não encontrada")
    return _calcular_kpis(auditoria)


@router.delete("/conciliacao/{auditoria_id}", status_code=204)
def excluir_auditoria(auditoria_id: int, db: Session = Depends(get_db)):
    auditoria = db.query(models.ConciliacaoAuditoria).filter(
        models.ConciliacaoAuditoria.id == auditoria_id
    ).first()
    if not auditoria:
        raise HTTPException(status_code=404, detail="Auditoria não encontrada")
    db.delete(auditoria)
    db.commit()


@router.get("/conciliacao/template/csv")
def baixar_template_csv():
    """Baixa um CSV de exemplo para preencher com dados do ERP."""
    conteudo = (
        "codigo_barra,nome,quantidade,estoque_minimo\n"
        "7891234567890,Produto Exemplo A,100,20\n"
        "7891234567891,Produto Exemplo B,50,10\n"
        "7891234567892,Produto Exemplo C,200,30\n"
    )
    return StreamingResponse(
        io.BytesIO(conteudo.encode("utf-8")),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=modelo_erp.csv"}
    )


@router.get("/conciliacao/{auditoria_id}/excel")
def exportar_excel_auditoria(auditoria_id: int, db: Session = Depends(get_db)):
    auditoria = db.query(models.ConciliacaoAuditoria).filter(
        models.ConciliacaoAuditoria.id == auditoria_id
    ).first()
    if not auditoria:
        raise HTTPException(status_code=404, detail="Auditoria não encontrada")

    kpis = _calcular_kpis(auditoria)
    itens = sorted(auditoria.itens,
                   key=lambda x: (x.status != "CRITICO", x.nome_produto))

    wb = openpyxl.Workbook()

    ws_resumo = wb.active
    ws_resumo.title = "Resumo KPIs"
    _excel_resumo(ws_resumo, auditoria, kpis)

    ws_itens = wb.create_sheet("Todos os Itens")
    _excel_itens(ws_itens, itens, "Todos os Itens")

    divergentes = [i for i in itens if i.status != "OK"]
    if divergentes:
        ws_div = wb.create_sheet("Divergências")
        _excel_itens(ws_div, divergentes, "Apenas Divergências")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    nome_arquivo = (
        f"auditoria_{auditoria_id}_"
        f"{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    )
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={nome_arquivo}"}
    )


# ── Excel helpers ─────────────────────────────────────────────────────────────

def _borda():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def _excel_resumo(ws, auditoria: models.ConciliacaoAuditoria, kpis: dict):
    COR_TITULO = "1F4E79"
    bd = _borda()

    # Título
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    nome = auditoria.nome or f"Auditoria #{auditoria.id}"
    c.value = f"AUDITORIA DE INVENTÁRIO — {nome.upper()}"
    c.font = Font(bold=True, size=14, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_TITULO)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo
    ws.merge_cells("A2:C2")
    c = ws["A2"]
    data = auditoria.criada_em.strftime("%d/%m/%Y %H:%M") if auditoria.criada_em else ""
    c.value = f"Emitido em: {data}  |  Sessão #{auditoria.sessao_id}"
    c.font = Font(italic=True, size=10, color="555555")
    c.alignment = Alignment(horizontal="center")

    metricas = [
        ("", ""),
        ("KPI", "Valor"),
        ("Acuracidade Geral", f"{kpis['acuracidade_geral']:.1f}%"),
        ("Total de Produtos Auditados", kpis["total_produtos"]),
        ("✅ Itens OK", f"{kpis['itens_ok']} ({kpis['taxa_ok']:.1f}%)"),
        ("⚠️  Divergentes (< 5%)", kpis["itens_divergentes"]),
        ("🚨 Críticos (≥ 5%)", kpis["itens_criticos"]),
        ("👻 Fantasmas (ERP sem físico)", kpis["itens_fantasma"]),
        ("🔵 Não Sistêmicos (físico sem ERP)", kpis["itens_nao_sistemico"]),
        ("🔴 Em Ruptura (abaixo do mínimo)", kpis["itens_em_ruptura"]),
        ("", ""),
        ("Total Físico (escaneado)", kpis["total_fisico"]),
        ("Total Sistêmico (ERP)", kpis["total_sistemico"]),
        ("Diferença Total (físico - sistêmico)", kpis["diferenca_total"]),
    ]

    for row_i, (kpi, val) in enumerate(metricas, start=3):
        bold = kpi in ("KPI", "Acuracidade Geral")
        c1 = ws.cell(row=row_i, column=1, value=kpi)
        c2 = ws.cell(row=row_i, column=2, value=val)
        c1.font = Font(bold=bold, size=11)
        c2.font = Font(bold=bold, size=11)
        if kpi == "KPI":
            c1.fill = PatternFill("solid", fgColor="D9D9D9")
            c2.fill = PatternFill("solid", fgColor="D9D9D9")
        if kpi:
            c1.border = bd
            c2.border = bd

    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 22


def _excel_itens(ws, itens, titulo: str):
    COR_CAB = "1F4E79"
    bd = _borda()

    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value = titulo.upper()
    c.font = Font(bold=True, size=12, color="FFFFFF")
    c.fill = PatternFill("solid", fgColor=COR_CAB)
    c.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    cabecalhos = [
        "Código de Barras", "Produto",
        "Físico", "Sistêmico", "Mín.",
        "Divergência", "Acurácia %", "Status"
    ]
    for col, cab in enumerate(cabecalhos, start=1):
        cell = ws.cell(row=3, column=col, value=cab)
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill("solid", fgColor=COR_CAB)
        cell.alignment = Alignment(horizontal="center")
        cell.border = bd

    for row_i, item in enumerate(itens, start=4):
        cor = STATUS_COR_EXCEL.get(item.status, "FFFFFF")
        fill = PatternFill("solid", fgColor=cor)
        linha = [
            item.codigo_barra,
            item.nome_produto,
            item.qtd_fisica,
            item.qtd_sistemica,
            item.estoque_minimo,
            item.divergencia,
            f"{item.acuracidade_item:.1f}%",
            STATUS_LABEL.get(item.status, item.status),
        ]
        for col, valor in enumerate(linha, start=1):
            cell = ws.cell(row=row_i, column=col, value=valor)
            cell.border = bd
            cell.fill = fill
            cell.alignment = Alignment(
                horizontal="center" if col in (3, 4, 5, 6, 7) else "left",
                vertical="center"
            )

    larguras = [20, 32, 9, 11, 7, 13, 12, 20]
    for col, larg in enumerate(larguras, start=1):
        ws.column_dimensions[get_column_letter(col)].width = larg

